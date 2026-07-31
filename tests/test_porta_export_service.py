from __future__ import annotations

import tempfile
import unittest
import hashlib
from pathlib import Path

from openpyxl import load_workbook

from services.porta_excel_export_service import EXPECTED_SHEETS, PortaExcelExportService
from services.porta_export_service import PortaExportService
from services.statuses import is_expense_eligible_for_accounting_export


PROJECT_ROOT = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = next(PROJECT_ROOT.glob("FORMATO*2026.xlsx"))


class FakeSheets:
    def __init__(self) -> None:
        self.companies = [{"company_id": "porta", "name": "Porta"}]
        self.cases = [
            {"case_id": "case-1", "company_id": "porta", "closed_at": "2026-06-30"},
            {"case_id": "case-2", "company_id": "porta", "closed_at": "2026-07-15"},
        ]
        self.expenses = [
            {
                "expense_id": "eligible",
                "case_id": "case-1",
                "status": "approved",
                "review_status": "rejected",
                "date": "2026-07-03",
                "cost_center": "Gastos de producción",
                "merchant": "Proveedor",
                "total_clp": 1000,
            },
            {
                "expense_id": "payment-is-irrelevant",
                "case_id": "case-1",
                "status": "pending_review",
                "review_status": "approved",
                "payment_status": "approved",
                "date": "2026-07-04",
                "cost_center": "Alimentación",
                "merchant": "Proveedor",
                "total_clp": 2000,
            },
            {
                "expense_id": "legacy-review",
                "case_id": "case-2",
                "status": "",
                "review_status": "approved",
                "date": "2026-07-05",
                "cost_center": "Facturas",
                "merchant": "Proveedor",
                "invoice_number": "F-1",
                "gross_amount": 1190,
            },
            {
                "expense_id": "incomplete-fee",
                "case_id": "case-2",
                "status": "approved",
                "date": "2026-07-06",
                "cost_center": "Boletas de honorarios",
                "merchant": "Prestador",
            },
        ]
        self.exports: dict[str, dict] = {}

    def list_companies(self): return self.companies
    def list_employees(self): return []
    def list_expense_cases(self): return self.cases
    def list_expenses(self): return self.expenses
    def create_porta_export(self, row): self.exports[row["export_id"]] = dict(row); return row
    def update_porta_export(self, export_id, payload):
        self.exports[export_id].update(payload)
        return self.exports[export_id]
    def get_porta_export(self, export_id): return self.exports.get(export_id)
    def list_porta_exports(self): return list(self.exports.values())


class FakeStorage:
    def __init__(self) -> None:
        self.objects: dict[str, bytes] = {}

    def upload_private_bytes(self, *, object_key, content, content_type):
        self.objects[object_key] = content

    def generate_signed_url(self, *, object_key, ttl_seconds):
        return f"https://example.test/{object_key}"


class PortaExportServiceTests(unittest.TestCase):
    def setUp(self) -> None:
        self.sheets = FakeSheets()
        self.storage = FakeStorage()
        self.service = PortaExportService(
            self.sheets,
            self.storage,
            PortaExcelExportService(str(TEMPLATE_PATH)),
        )
        self.user = {
            "email": "admin@porta.cl",
            "role": "company_admin",
            "scope_type": "company",
            "company_id": "porta",
            "company_ids": ["porta"],
        }

    def test_operational_status_is_authoritative_and_payment_is_ignored(self):
        self.assertTrue(is_expense_eligible_for_accounting_export(self.sheets.expenses[0]))
        self.assertFalse(is_expense_eligible_for_accounting_export(self.sheets.expenses[1]))
        self.assertTrue(is_expense_eligible_for_accounting_export(self.sheets.expenses[2]))

    def test_month_uses_document_date_by_default_and_marks_incomplete_expenses(self):
        preview = self.service.preview(
            user=self.user, company_id="porta", scope="month", year=2026, month=7
        )
        self.assertEqual(preview["date_source"], "document_date")
        self.assertEqual(preview["included_count"], 2)
        self.assertEqual(preview["excluded_count"], 1)
        self.assertIn("Número de documento incompleto", preview["excluded"][0]["reasons"])
        self.assertEqual(preview["by_sheet"]["Gastos de Producción"], 1)
        self.assertEqual(preview["by_sheet"]["Facturas"], 1)

    def test_case_closed_at_rule_is_configurable(self):
        preview = self.service.preview(
            user=self.user,
            company_id="porta",
            scope="month",
            year=2026,
            month=6,
            date_source="case_closed_at",
        )
        self.assertEqual(preview["included_count"], 1)

    def test_each_generation_creates_an_independent_history_row(self):
        first = self.service.generate(
            user=self.user, company_id="porta", scope="case", case_id="case-1"
        )
        second = self.service.generate(
            user=self.user, company_id="porta", scope="case", case_id="case-1"
        )
        self.assertNotEqual(first["export_id"], second["export_id"])
        self.assertEqual(len(self.sheets.exports), 2)
        self.assertEqual(len(self.storage.objects), 2)


class PortaWorkbookFidelityTests(unittest.TestCase):
    def test_dynamic_rows_keep_template_structure_styles_and_summary_reference(self):
        renderer = PortaExcelExportService(str(TEMPLATE_PATH))
        template_hash = hashlib.sha256(TEMPLATE_PATH.read_bytes()).hexdigest()
        expenses = [
            {
                "porta_sheet": "Gastos de Producción",
                "date": "2026-07-01",
                "document_number": str(index),
                "detail": "Gasto",
                "gross_amount": 1000,
                "net_amount": 0,
                "tax_amount": 0,
                "withholding_amount": 0,
            }
            for index in range(35)
        ]
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as output:
            output.write(renderer.render(expenses))
            output.flush()
            original = load_workbook(TEMPLATE_PATH, data_only=False)
            generated = load_workbook(output.name, data_only=False)

        self.assertEqual(generated.sheetnames, EXPECTED_SHEETS)
        for sheet_name in EXPECTED_SHEETS:
            self.assertEqual(
                list(generated[sheet_name].merged_cells.ranges),
                list(original[sheet_name].merged_cells.ranges),
            )
        production = generated["Gastos de Producción"]
        self.assertEqual(production["D47"].value, "=SUM(D12:D46)")
        self.assertEqual(generated["Resumen"]["C12"].value, "=+'Gastos de Producción'!D47")
        self.assertEqual(production["A46"].style_id, production["A12"].style_id)
        self.assertEqual(production["D46"].number_format, production["D12"].number_format)
        self.assertEqual(
            [generated["Resumen"][cell].value for cell in ("C4", "C5", "C6", "C7", "C9")],
            [None, None, None, None, None],
        )
        self.assertEqual(hashlib.sha256(TEMPLATE_PATH.read_bytes()).hexdigest(), template_hash)


if __name__ == "__main__":
    unittest.main()
