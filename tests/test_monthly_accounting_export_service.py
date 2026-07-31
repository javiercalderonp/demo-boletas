from __future__ import annotations

import io
import json
import unittest
import zipfile

from openpyxl import load_workbook

from services.monthly_accounting_export_service import MonthlyAccountingExportService


class FakeSheets:
    def __init__(self) -> None:
        self.exports: list[dict] = []
        self.companies = [
            {"company_id": "acme", "name": "Ácme / Chile", "rut": "76.000.000-1", "active": True},
            {"company_id": "other", "name": "Otra", "active": True},
        ]
        self.employees = [
            {"phone": "+56911111111", "name": "Ana Pérez", "rut": "11.111.111-1", "email": "ana@example.com", "company_id": "acme"},
            {"phone": "+56922222222", "name": "Otro", "company_id": "other"},
        ]
        self.cases = [
            {"case_id": "CASE/1", "company_id": "acme", "employee_phone": "+56911111111", "context_label": "../Viaje Norte", "created_at": "2026-06-01T00:00:00Z", "status": "closed", "rendicion_status": "closed", "fondos_entregados": "100000", "settlement_amount_clp": "20000"},
            {"case_id": "CASE-OTHER", "company_id": "other", "employee_phone": "+56922222222", "created_at": "2026-07-01T00:00:00Z"},
        ]
        self.expenses = [
            {"expense_id": "EXP-1", "case_id": "CASE/1", "date": "2026-07-10", "created_at": "2026-07-10T10:00:00Z", "merchant": "Café Uno", "currency": "CLP", "total": "120000", "total_clp": "120000", "status": "approved", "cost_center": "Ventas", "receipt_object_key": "receipts/a/duplicado.jpg"},
            {"expense_id": "EXP-2", "case_id": "CASE/1", "date": "2026-07-11", "merchant": "Café Uno", "currency": "USD", "total": "10", "total_clp": "9500", "status": "pending", "receipt_object_key": "receipts/b/duplicado.jpg"},
            {"expense_id": "EXP-X", "case_id": "CASE-OTHER", "date": "2026-07-10", "total_clp": "999999", "status": "approved"},
        ]
        self.documents = [{"document_id": "DOC-1", "case_id": "CASE/1", "object_key": "reports/case.pdf", "signature_status": "completed"}]

    def list_companies(self):
        return list(self.companies)

    def list_employees(self):
        return list(self.employees)

    def list_expense_cases(self):
        return list(self.cases)

    def list_expenses(self):
        return list(self.expenses)

    def list_expense_case_documents_by_phone_case(self, phone, case_id):
        return [row for row in self.documents if row["case_id"] == case_id]

    def create_monthly_accounting_export(self, payload):
        self.exports.append(dict(payload))
        return payload

    def update_monthly_accounting_export(self, export_id, payload):
        row = next(item for item in self.exports if item["export_id"] == export_id)
        row.update(payload)
        return dict(row)

    def get_monthly_accounting_export(self, export_id):
        return next((dict(item) for item in self.exports if item["export_id"] == export_id), None)

    def list_monthly_accounting_exports(self):
        return [dict(item) for item in reversed(self.exports)]

    def delete_monthly_accounting_export(self, export_id):
        row = next((item for item in self.exports if item["export_id"] == export_id), None)
        if row is None:
            return None
        self.exports.remove(row)
        return dict(row)


class FakeStorage:
    def __init__(self, *, missing_receipt: bool = False) -> None:
        self.uploaded: dict[str, bytes] = {}
        self.missing_receipt = missing_receipt

    def upload_private_bytes(self, *, object_key, content, content_type):
        self.uploaded[object_key] = content
        return {"object_key": object_key}

    def download_private_bytes(self, *, object_key):
        if self.missing_receipt and object_key.startswith("receipts/"):
            raise RuntimeError("GCS inaccessible")
        return b"%PDF-test" if object_key.endswith(".pdf") else b"image"

    def generate_signed_url(self, *, object_key, ttl_seconds):
        return f"https://signed.example/{object_key}?ttl={ttl_seconds}"

    def delete_private_object(self, *, object_key):
        self.uploaded.pop(object_key, None)


class MonthlyAccountingExportServiceTests(unittest.TestCase):
    def setUp(self):
        self.sheets = FakeSheets()
        self.storage = FakeStorage()
        self.service = MonthlyAccountingExportService(self.sheets, self.storage)
        self.global_user = {"email": "admin@example.com", "name": "Admin", "role": "super_admin", "scope_type": "global"}
        self.scoped_user = {"email": "acme@example.com", "role": "company_admin", "scope_type": "company", "company_ids": ["acme"]}

    def test_preview_filters_month_and_company_and_uses_decimal_totals(self):
        preview = self.service.preview(user=self.scoped_user, company_id="acme", year=2026, month=7)
        self.assertEqual(preview["case_count"], 1)
        self.assertEqual(preview["expense_count"], 2)
        self.assertEqual(preview["total_clp"], 129500)
        self.assertEqual(preview["approved_count"], 1)
        self.assertEqual(preview["pending_count"], 1)
        self.assertEqual(preview["without_cost_center_count"], 1)

    def test_user_cannot_export_another_company(self):
        with self.assertRaises(PermissionError):
            self.service.preview(user=self.scoped_user, company_id="other", year=2026, month=7)

    def test_generate_builds_private_files_manifest_zip_and_excel_numbers(self):
        result = self.service.generate(user=self.global_user, company_id="acme", year=2026, month=7)
        self.assertEqual(result["status"], "completed_with_warnings")
        zip_key = next(key for key in self.storage.uploaded if key.endswith(".zip"))
        with zipfile.ZipFile(io.BytesIO(self.storage.uploaded[zip_key])) as archive:
            names = archive.namelist()
            self.assertIn("manifest.json", names)
            self.assertFalse(any(".." in name for name in names))
            manifest = json.loads(archive.read("manifest.json"))
            self.assertEqual(manifest["numero_gastos"], 2)
            self.assertEqual(manifest["total_clp"], 129500)
            receipt_names = [name for name in names if "/comprobantes/" in name]
            self.assertEqual(len(receipt_names), 2)
            self.assertEqual(len(set(receipt_names)), 2)
        xlsx_key = next(key for key in self.storage.uploaded if key.endswith(".xlsx"))
        workbook = load_workbook(io.BytesIO(self.storage.uploaded[xlsx_key]))
        self.assertEqual(
            workbook.sheetnames,
            ["Resumen", "Gastos", "Rendiciones", "Personas", "Centros de costo", "Excepciones"],
        )
        self.assertIsInstance(workbook["Gastos"]["U2"].value, (int, float))
        self.assertEqual(workbook["Gastos"].freeze_panes, "A2")

    def test_missing_gcs_receipt_finishes_with_warning(self):
        service = MonthlyAccountingExportService(self.sheets, FakeStorage(missing_receipt=True))
        result = service.generate(user=self.global_user, company_id="acme", year=2026, month=7)
        warning_types = {warning["type"] for warning in result["warnings"]}
        self.assertIn("receipt_unavailable", warning_types)
        self.assertEqual(result["status"], "completed_with_warnings")

    def test_empty_month_is_exportable(self):
        preview = self.service.preview(user=self.global_user, company_id="acme", year=2025, month=1)
        self.assertEqual(preview["expense_count"], 0)
        self.assertEqual(preview["total_clp"], 0)

    def test_preview_supports_custom_date_range(self):
        preview = self.service.preview(
            user=self.scoped_user,
            company_id="acme",
            year=2026,
            month=7,
            date_from="2026-07-11",
            date_to="2026-07-11",
        )
        self.assertEqual(preview["period"], "2026-07-11 al 2026-07-11")
        self.assertEqual(preview["case_count"], 1)
        self.assertEqual(preview["expense_count"], 1)
        self.assertEqual(preview["total_clp"], 9500)

    def test_custom_date_range_must_be_ordered(self):
        with self.assertRaises(ValueError):
            self.service.preview(
                user=self.scoped_user,
                company_id="acme",
                year=2026,
                month=7,
                date_from="2026-07-12",
                date_to="2026-07-10",
            )

    def test_closed_case_in_period_includes_all_its_expenses(self):
        self.sheets.cases[0]["created_at"] = "2026-05-01T00:00:00Z"
        self.sheets.cases[0]["closed_at"] = "2026-06-30T12:00:00Z"
        self.sheets.expenses[0]["date"] = "2026-05-10"
        self.sheets.expenses[1]["date"] = "2026-05-11"

        preview = self.service.preview(
            user=self.scoped_user,
            company_id="acme",
            year=2026,
            month=6,
        )

        self.assertEqual(preview["case_count"], 1)
        self.assertEqual(preview["expense_count"], 2)
        self.assertEqual(preview["total_clp"], 129500)

    def test_signed_url_is_short_lived_and_scoped(self):
        result = self.service.generate(user=self.global_user, company_id="acme", year=2026, month=7)
        url = self.service.signed_download_url(self.scoped_user, result["export_id"], "pdf")
        self.assertIn("ttl=300", url)
        denied = {"role": "company_admin", "scope_type": "company", "company_ids": ["other"]}
        self.assertIsNone(self.service.signed_download_url(denied, result["export_id"], "pdf"))

    def test_delete_removes_export_and_generated_files(self):
        result = self.service.generate(
            user=self.global_user, company_id="acme", year=2026, month=7
        )
        self.assertTrue(self.storage.uploaded)

        deleted = self.service.delete_for_user(self.scoped_user, result["export_id"])

        self.assertEqual(deleted["export_id"], result["export_id"])
        self.assertFalse(self.storage.uploaded)
        self.assertIsNone(self.sheets.get_monthly_accounting_export(result["export_id"]))

    def test_delete_is_scoped_to_accessible_companies(self):
        result = self.service.generate(
            user=self.global_user, company_id="acme", year=2026, month=7
        )
        denied = {"role": "company_admin", "scope_type": "company", "company_ids": ["other"]}

        self.assertIsNone(self.service.delete_for_user(denied, result["export_id"]))
        self.assertIsNotNone(self.sheets.get_monthly_accounting_export(result["export_id"]))


if __name__ == "__main__":
    unittest.main()
