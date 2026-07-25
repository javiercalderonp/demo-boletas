from __future__ import annotations

import csv
import io
import json
import logging
import re
import unicodedata
import zipfile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import PurePosixPath
from typing import Any

from services.backoffice_permissions import can_access_company
from services.sheets_service import SheetsService, normalize_cost_centers
from services.storage_service import GCSStorageService
from utils.helpers import json_dumps, json_loads, make_id, utc_now_iso

logger = logging.getLogger(__name__)
FORMAT_VERSION = "1.0"
MAX_EXPENSES = 5000
MAX_ZIP_BYTES = 250 * 1024 * 1024
COMPLETED_STATUSES = {"completed", "completed_with_warnings"}


def _decimal(value: Any) -> Decimal:
    text = str(value if value not in (None, "") else "0").strip()
    try:
        return Decimal(text.replace(".", "").replace(",", ".") if "," in text else text)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _clp(value: Decimal) -> int:
    return int(value.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _safe_name(value: Any, fallback: str = "archivo") -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode()
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", text).strip("._")
    return text[:100] or fallback


def _date_in_period(value: Any, year: int, month: int) -> bool:
    text = str(value or "").strip()
    if len(text) < 7:
        return False
    return text[:7] == f"{year:04d}-{month:02d}"


def _status(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


@dataclass
class MonthlyAccountingExportService:
    sheets_service: SheetsService
    storage_service: GCSStorageService

    def preview(
        self,
        *,
        user: dict[str, Any],
        company_id: str,
        year: int,
        month: int,
        cost_center: str = "",
        case_status: str = "",
        expense_status: str = "",
    ) -> dict[str, Any]:
        dataset = self._dataset(
            user=user,
            company_id=company_id,
            year=year,
            month=month,
            cost_center=cost_center,
            case_status=case_status,
            expense_status=expense_status,
        )
        return self._summary(dataset)

    def generate(
        self,
        *,
        user: dict[str, Any],
        company_id: str,
        year: int,
        month: int,
        cost_center: str = "",
        case_status: str = "",
        expense_status: str = "",
        include_csv: bool = True,
    ) -> dict[str, Any]:
        dataset = self._dataset(
            user=user,
            company_id=company_id,
            year=year,
            month=month,
            cost_center=cost_center,
            case_status=case_status,
            expense_status=expense_status,
        )
        if len(dataset["expenses"]) > MAX_EXPENSES:
            raise ValueError(f"La exportación supera el límite de {MAX_EXPENSES} gastos")
        export_id = make_id("acct")
        now = utc_now_iso()
        filters = {
            "cost_center": cost_center,
            "case_status": case_status,
            "expense_status": expense_status,
            "include_csv": include_csv,
        }
        row = {
            "export_id": export_id,
            "company_id": company_id,
            "period_year": year,
            "period_month": month,
            "filters_json": json_dumps(filters),
            "status": "processing",
            "requested_by": str(user.get("email", "") or ""),
            "requested_at": now,
            "completed_at": "",
            "expense_count": len(dataset["expenses"]),
            "case_count": len(dataset["cases"]),
            "employee_count": len(dataset["employees"]),
            "total_clp": self._summary(dataset)["total_clp"],
            "warnings_json": "[]",
            "snapshot_json": json_dumps(self._snapshot(dataset)),
            "format_version": FORMAT_VERSION,
        }
        self.sheets_service.create_monthly_accounting_export(row)
        try:
            summary = self._summary(dataset)
            warnings = list(summary["warnings"])
            pdf_bytes = self._pdf(dataset, summary, export_id, user, now)
            xlsx_bytes = self._xlsx(dataset, summary, warnings, export_id)
            csv_bytes = self._csv(dataset) if include_csv else b""
            zip_bytes, zip_warnings, manifest = self._zip(
                dataset, summary, export_id, user, now, pdf_bytes, xlsx_bytes, csv_bytes
            )
            warnings.extend(zip_warnings)
            company_slug = _safe_name(dataset["company"].get("name") or company_id)
            period = f"{year:04d}-{month:02d}"
            prefix = f"accounting-exports/{company_slug}/{period}/{export_id}"
            objects = {
                "pdf_object_key": (f"{prefix}/Informe_Gastos_{company_slug}_{period}.pdf", pdf_bytes, "application/pdf"),
                "xlsx_object_key": (f"{prefix}/Detalle_Gastos_{company_slug}_{period}.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                "zip_object_key": (f"{prefix}/Cierre_Contable_{company_slug}_{period}.zip", zip_bytes, "application/zip"),
                "manifest_object_key": (f"{prefix}/manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2).encode(), "application/json"),
            }
            if csv_bytes:
                objects["csv_object_key"] = (
                    f"{prefix}/Gastos_{company_slug}_{period}.csv",
                    csv_bytes,
                    "text/csv; charset=utf-8",
                )
            keys: dict[str, str] = {}
            for field, (key, content, content_type) in objects.items():
                self.storage_service.upload_private_bytes(
                    object_key=key, content=content, content_type=content_type
                )
                keys[field] = key
            result_status = "completed_with_warnings" if warnings else "completed"
            updated = self.sheets_service.update_monthly_accounting_export(
                export_id,
                {
                    **keys,
                    "status": result_status,
                    "completed_at": utc_now_iso(),
                    "warnings_json": json_dumps(warnings),
                },
            )
            return self._public_row(updated or row)
        except Exception as exc:
            self.sheets_service.update_monthly_accounting_export(
                export_id, {"status": "failed", "error_message": str(exc)}
            )
            logger.exception("Monthly accounting export failed export_id=%s", export_id)
            raise

    def list_for_user(self, user: dict[str, Any]) -> list[dict[str, Any]]:
        rows = self.sheets_service.list_monthly_accounting_exports()
        return [
            self._public_row(row)
            for row in rows
            if can_access_company(user, row.get("company_id", ""))
        ][:20]

    def get_for_user(self, user: dict[str, Any], export_id: str) -> dict[str, Any] | None:
        row = self.sheets_service.get_monthly_accounting_export(export_id)
        if not row or not can_access_company(user, row.get("company_id", "")):
            return None
        return self._public_row(row)

    def delete_for_user(self, user: dict[str, Any], export_id: str) -> dict[str, Any] | None:
        row = self.sheets_service.get_monthly_accounting_export(export_id)
        if not row or not can_access_company(user, row.get("company_id", "")):
            return None

        object_fields = (
            "pdf_object_key",
            "xlsx_object_key",
            "csv_object_key",
            "zip_object_key",
            "manifest_object_key",
        )
        for field in object_fields:
            object_key = str(row.get(field, "") or "").strip()
            if not object_key:
                continue
            self.storage_service.delete_private_object(object_key=object_key)

        deleted = self.sheets_service.delete_monthly_accounting_export(export_id)
        return self._public_row(deleted) if deleted else None

    def signed_download_url(
        self, user: dict[str, Any], export_id: str, file_format: str
    ) -> str | None:
        row = self.sheets_service.get_monthly_accounting_export(export_id)
        if not row or not can_access_company(user, row.get("company_id", "")):
            return None
        field = {"pdf": "pdf_object_key", "xlsx": "xlsx_object_key", "csv": "csv_object_key", "zip": "zip_object_key"}.get(file_format)
        if not field or not row.get(field):
            return None
        return self.storage_service.generate_signed_url(
            object_key=str(row[field]), ttl_seconds=300
        )

    def _dataset(self, *, user: dict[str, Any], company_id: str, year: int, month: int, cost_center: str, case_status: str, expense_status: str) -> dict[str, Any]:
        if year < 2000 or year > 2100 or month < 1 or month > 12:
            raise ValueError("Periodo inválido")
        if not company_id or not can_access_company(user, company_id):
            raise PermissionError("No tienes acceso a la empresa seleccionada")
        companies = self.sheets_service.list_companies()
        company = next((c for c in companies if str(c.get("company_id")) == company_id), None)
        if company is None:
            raise ValueError("Empresa inexistente")
        employees = [e for e in self.sheets_service.list_employees() if str(e.get("company_id")) == company_id]
        employees_by_phone = {str(e.get("phone", "")): e for e in employees}
        company_cases = []
        for case in self.sheets_service.list_expense_cases():
            phone = str(case.get("employee_phone", case.get("phone", "")) or "")
            resolved_company = str(case.get("company_id", "") or (employees_by_phone.get(phone) or {}).get("company_id", ""))
            if resolved_company != company_id:
                continue
            if case_status and _status(case.get("rendicion_status") or case.get("status")) != _status(case_status):
                continue
            company_cases.append(case)
        company_cases_by_id = {str(c.get("case_id", "")): c for c in company_cases}
        expenses = []
        for expense in self.sheets_service.list_expenses():
            if str(expense.get("case_id", "")) not in company_cases_by_id:
                continue
            expense_date = expense.get("date") or expense.get("created_at")
            if not _date_in_period(expense_date, year, month):
                continue
            if cost_center and str(expense.get("cost_center", "")).strip().lower() != cost_center.strip().lower():
                continue
            if expense_status and _status(expense.get("status")) != _status(expense_status):
                continue
            expenses.append(expense)
        included_case_ids = {str(e.get("case_id", "")) for e in expenses}
        cases = [
            case
            for case in company_cases
            if str(case.get("case_id", "")) in included_case_ids
            or (
                not cost_center
                and not expense_status
                and _date_in_period(
                    case.get("created_at") or case.get("closed_at") or case.get("updated_at"),
                    year,
                    month,
                )
            )
        ]
        cases_by_id = {str(c.get("case_id", "")): c for c in cases}
        phones = {str(c.get("employee_phone", c.get("phone", "")) or "") for c in cases}
        included_employees = [e for e in employees if str(e.get("phone", "")) in phones]
        documents = []
        for case in cases:
            phone = str(case.get("employee_phone", case.get("phone", "")) or "")
            documents.extend(self.sheets_service.list_expense_case_documents_by_phone_case(phone, str(case.get("case_id", ""))))
        return {"company": company, "employees": included_employees, "employees_by_phone": employees_by_phone, "cases": cases, "cases_by_id": cases_by_id, "expenses": expenses, "documents": documents, "year": year, "month": month}

    def _summary(self, data: dict[str, Any]) -> dict[str, Any]:
        amounts = {"approved": Decimal("0"), "pending": Decimal("0"), "rejected": Decimal("0"), "observed": Decimal("0"), "processing": Decimal("0")}
        counts = {key: 0 for key in amounts}
        warnings: list[dict[str, str]] = []
        total = Decimal("0")
        without_receipt = 0
        without_center = 0
        for expense in data["expenses"]:
            amount = _decimal(expense.get("total_clp"))
            total += amount
            state = _status(expense.get("status") or expense.get("processing_status"))
            bucket = "approved" if state in {"approved", "aprobado"} else "rejected" if state in {"rejected", "rechazado"} else "observed" if state in {"observed", "observado", "needs_review"} else "processing" if state in {"processing", "procesando"} else "pending"
            amounts[bucket] += amount
            counts[bucket] += 1
            if not str(expense.get("cost_center", "") or "").strip():
                without_center += 1
            if not str(expense.get("receipt_object_key", "") or "").strip():
                without_receipt += 1
        open_cases = sum(1 for c in data["cases"] if _status(c.get("rendicion_status") or c.get("status")) in {"open", "active", "pending_user_confirmation"})
        if open_cases:
            warnings.append({"type": "open_cases", "description": f"{open_cases} rendiciones aún están abiertas", "severity": "warning"})
        if counts["pending"]:
            warnings.append({"type": "pending_expenses", "description": f"{counts['pending']} gastos están pendientes", "severity": "warning"})
        if without_receipt:
            warnings.append({"type": "missing_receipt", "description": f"{without_receipt} gastos no tienen comprobante almacenado", "severity": "warning"})
        if without_center:
            warnings.append({"type": "missing_cost_center", "description": f"{without_center} gastos no tienen centro de costo", "severity": "warning"})
        advances = sum((_decimal(c.get("fondos_entregados")) for c in data["cases"]), Decimal("0"))
        return {
            "period": f"{data['year']:04d}-{data['month']:02d}",
            "company_id": data["company"].get("company_id"),
            "company_name": data["company"].get("name"),
            "case_count": len(data["cases"]),
            "expense_count": len(data["expenses"]),
            "employee_count": len(data["employees"]),
            "total_clp": _clp(total),
            "approved_count": counts["approved"],
            "pending_count": counts["pending"],
            "rejected_count": counts["rejected"],
            "observed_count": counts["observed"],
            "amounts": {key: _clp(value) for key, value in amounts.items()},
            "advances_clp": _clp(advances),
            "without_receipt_count": without_receipt,
            "without_cost_center_count": without_center,
            "open_case_count": open_cases,
            "warnings": warnings,
        }

    def _snapshot(self, data: dict[str, Any]) -> dict[str, Any]:
        return {
            "expense_ids": [e.get("expense_id") for e in data["expenses"]],
            "case_ids": [c.get("case_id") for c in data["cases"]],
            "employee_phones": [e.get("phone") for e in data["employees"]],
            "generated_from_updated_at": max([str(e.get("updated_at", "") or "") for e in data["expenses"]] or [""]),
        }

    def _expense_rows(self, data: dict[str, Any]) -> list[list[Any]]:
        rows: list[list[Any]] = []
        for expense in data["expenses"]:
            case = data["cases_by_id"].get(str(expense.get("case_id", "")), {})
            phone = str(case.get("employee_phone", case.get("phone", expense.get("phone", ""))) or "")
            employee = data["employees_by_phone"].get(phone, {})
            rows.append([
                f"{data['year']:04d}-{data['month']:02d}", expense.get("expense_id", ""), expense.get("case_id", ""),
                case.get("context_label", ""), case.get("rendicion_status", case.get("status", "")),
                employee.get("name", ""), employee.get("rut", ""), employee.get("email", ""), employee.get("phone", ""),
                data["company"].get("name", ""), expense.get("date", ""), expense.get("created_at", ""),
                expense.get("merchant", ""), expense.get("issuer_tax_id", ""), expense.get("document_type", ""),
                expense.get("invoice_number", ""), expense.get("category", ""), expense.get("cost_center", ""),
                expense.get("currency", ""), float(_decimal(expense.get("total"))), float(_decimal(expense.get("total_clp"))),
                float(_decimal(expense.get("net_amount"))), float(_decimal(expense.get("tax_amount"))), float(_decimal(expense.get("total_clp"))),
                expense.get("status", ""), expense.get("review_reason", ""), expense.get("updated_at", ""),
                bool(expense.get("receipt_object_key")), PurePosixPath(str(expense.get("receipt_object_key", ""))).name,
                self._receipt_zip_path(expense, case, employee),
            ])
        return rows

    @staticmethod
    def _expense_headers() -> list[str]:
        return ["periodo", "expense_id", "case_id", "nombre_rendicion", "estado_rendicion", "persona", "rut_persona", "email", "telefono", "empresa", "fecha_gasto", "fecha_recepcion", "proveedor", "rut_emisor", "tipo_documento", "folio", "categoria", "centro_costo", "moneda", "monto_original", "monto_clp", "monto_neto", "iva", "total", "estado", "motivo_revision", "fecha_aprobacion", "comprobante_disponible", "archivo_comprobante", "ruta_zip"]

    def _pdf(self, data: dict[str, Any], summary: dict[str, Any], export_id: str, user: dict[str, Any], generated_at: str) -> bytes:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        except ImportError as exc:
            raise RuntimeError("Falta reportlab") from exc
        buffer = io.BytesIO()
        styles = getSampleStyleSheet()
        company_name = str(data["company"].get("name") or data["company"].get("company_id"))
        period = summary["period"]
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=14*mm, bottomMargin=15*mm, title=f"Informe mensual {period}")
        story: list[Any] = [
            Paragraph(company_name, styles["Title"]),
            Spacer(1, 6),
            Paragraph("Informe mensual de rendiciones y gastos", styles["Heading1"]),
            Paragraph(f"Periodo: {period} · Generado: {generated_at} · Usuario: {user.get('name') or user.get('email')} · Exportación: {export_id}", styles["BodyText"]),
            Spacer(1, 12),
            Paragraph("Resumen ejecutivo", styles["Heading2"]),
        ]
        summary_rows = [
            ["Personas", summary["employee_count"], "Rendiciones", summary["case_count"], "Gastos", summary["expense_count"]],
            ["Total rendido", self._format_clp(summary["total_clp"]), "Aprobado", self._format_clp(summary["amounts"]["approved"]), "Pendiente", self._format_clp(summary["amounts"]["pending"])],
            ["Rechazado", self._format_clp(summary["amounts"]["rejected"]), "Anticipos", self._format_clp(summary["advances_clp"]), "Sin comprobante", summary["without_receipt_count"]],
        ]
        story.append(self._pdf_table(summary_rows, Table, TableStyle, colors))
        story.extend([Spacer(1, 10), Paragraph("Resumen por estado", styles["Heading2"])])
        state_rows = [["Estado", "Gastos", "Monto", "% total"]]
        for key, label in [("approved", "Aprobado"), ("pending", "Pendiente"), ("rejected", "Rechazado"), ("observed", "Observado"), ("processing", "En procesamiento")]:
            pct = (summary["amounts"][key] / summary["total_clp"] * 100) if summary["total_clp"] else 0
            state_rows.append([label, summary[f"{key}_count"] if f"{key}_count" in summary else 0, self._format_clp(summary["amounts"][key]), f"{pct:.1f}%"])
        story.append(self._pdf_table(state_rows, Table, TableStyle, colors))
        story.extend([PageBreak(), Paragraph("Detalle de gastos", styles["Heading2"])])
        detail = [["N°", "Fecha", "Persona", "Rendición", "Centro", "Proveedor", "Documento", "Monto CLP", "Estado", "Comprobante"]]
        for index, expense in enumerate(data["expenses"], 1):
            case = data["cases_by_id"].get(str(expense.get("case_id", "")), {})
            employee = data["employees_by_phone"].get(str(case.get("employee_phone", case.get("phone", "")) or ""), {})
            detail.append([index, expense.get("date", ""), employee.get("name", ""), case.get("context_label", ""), expense.get("cost_center") or "Sin centro", expense.get("merchant", ""), expense.get("document_type", ""), self._format_clp(_clp(_decimal(expense.get("total_clp")))), expense.get("status", ""), "Sí" if expense.get("receipt_object_key") else "No"])
        story.append(self._pdf_table(detail, Table, TableStyle, colors, repeat_rows=1, font_size=6))
        if summary["warnings"]:
            story.extend([PageBreak(), Paragraph("Excepciones y observaciones", styles["Heading2"])])
            for warning in summary["warnings"]:
                story.append(Paragraph(f"• {warning['description']}", styles["BodyText"]))
        disclaimer = "Resumen administrativo. Los documentos originales deben conservarse como respaldo; este informe no reemplaza libros contables, declaraciones, DTE ni el Registro de Compras y Ventas."
        def footer(canvas: Any, current_doc: Any) -> None:
            canvas.saveState()
            canvas.setFont("Helvetica", 7)
            canvas.drawString(12*mm, 7*mm, f"{company_name} · {period} · {export_id}")
            canvas.drawRightString(285*mm, 7*mm, f"Página {current_doc.page}")
            canvas.restoreState()
        story.extend([Spacer(1, 8), Paragraph(disclaimer, styles["Italic"])])
        doc.build(story, onFirstPage=footer, onLaterPages=footer)
        return buffer.getvalue()

    @staticmethod
    def _pdf_table(rows: list[list[Any]], table_cls: Any, style_cls: Any, colors: Any, repeat_rows: int = 0, font_size: int = 8) -> Any:
        table = table_cls(rows, repeatRows=repeat_rows, hAlign="LEFT")
        table.setStyle(style_cls([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#e5e7eb")), ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#9ca3af")), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"), ("FONTSIZE", (0,0), (-1,-1), font_size), ("VALIGN", (0,0), (-1,-1), "TOP"), ("LEFTPADDING", (0,0), (-1,-1), 3), ("RIGHTPADDING", (0,0), (-1,-1), 3)]))
        return table

    def _xlsx(self, data: dict[str, Any], summary: dict[str, Any], warnings: list[dict[str, str]], export_id: str) -> bytes:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError as exc:
            raise RuntimeError("Falta openpyxl") from exc
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen"
        summary_values = [["Empresa", summary["company_name"]], ["Periodo", summary["period"]], ["Exportación", export_id], ["Total rendiciones", summary["case_count"]], ["Total gastos", summary["expense_count"]], ["Total rendido", summary["total_clp"]], ["Total aprobado", summary["amounts"]["approved"]], ["Total pendiente", summary["amounts"]["pending"]], ["Total rechazado", summary["amounts"]["rejected"]], ["Anticipos entregados", summary["advances_clp"]], ["Sin centro de costo", summary["without_cost_center_count"]], ["Sin comprobante", summary["without_receipt_count"]]]
        for row in summary_values:
            ws.append(row)
        expenses_ws = wb.create_sheet("Gastos")
        expenses_ws.append(self._expense_headers())
        for row in self._expense_rows(data):
            expenses_ws.append(row)
        cases_ws = wb.create_sheet("Rendiciones")
        cases_ws.append(["case_id", "nombre", "responsable", "rut", "estado", "fecha_creacion", "fecha_cierre", "fondos_entregados", "total_rendido", "total_aprobado", "total_rechazado", "direccion_liquidacion", "monto_liquidacion", "cantidad_gastos", "centros_costo", "pdf_consolidado", "estado_firma"])
        documents_by_case = {str(d.get("case_id", "")): d for d in data["documents"] if d.get("object_key")}
        for case in data["cases"]:
            case_id = str(case.get("case_id", ""))
            employee = data["employees_by_phone"].get(str(case.get("employee_phone", case.get("phone", "")) or ""), {})
            case_expenses = [e for e in data["expenses"] if str(e.get("case_id", "")) == case_id]
            approved = sum((_decimal(e.get("total_clp")) for e in case_expenses if _status(e.get("status")) == "approved"), Decimal("0"))
            rejected = sum((_decimal(e.get("total_clp")) for e in case_expenses if _status(e.get("status")) == "rejected"), Decimal("0"))
            doc = documents_by_case.get(case_id, {})
            cases_ws.append([case_id, case.get("context_label", ""), employee.get("name", ""), employee.get("rut", ""), case.get("rendicion_status", case.get("status", "")), case.get("created_at", ""), case.get("closed_at", ""), float(_decimal(case.get("fondos_entregados"))), float(sum((_decimal(e.get("total_clp")) for e in case_expenses), Decimal("0"))), float(approved), float(rejected), case.get("settlement_direction", ""), float(_decimal(case.get("settlement_amount_clp"))), len(case_expenses), ", ".join(normalize_cost_centers(case.get("cost_centers"))), bool(doc), doc.get("signature_status", "")])
        people_ws = wb.create_sheet("Personas")
        people_ws.append(["employee_id", "nombre", "apellido", "rut", "email", "telefono", "empresa", "rendiciones", "gastos", "total_rendido", "total_aprobado", "saldo_acumulado"])
        for employee in data["employees"]:
            phone = str(employee.get("phone", ""))
            employee_cases = [c for c in data["cases"] if str(c.get("employee_phone", c.get("phone", "")) or "") == phone]
            ids = {str(c.get("case_id", "")) for c in employee_cases}
            employee_expenses = [e for e in data["expenses"] if str(e.get("case_id", "")) in ids]
            rendered = sum((_decimal(e.get("total_clp")) for e in employee_expenses), Decimal("0"))
            approved = sum((_decimal(e.get("total_clp")) for e in employee_expenses if _status(e.get("status")) == "approved"), Decimal("0"))
            people_ws.append([phone, employee.get("name", ""), employee.get("last_name", ""), employee.get("rut", ""), employee.get("email", ""), phone, summary["company_name"], len(employee_cases), len(employee_expenses), float(rendered), float(approved), float(approved - sum((_decimal(c.get("fondos_entregados")) for c in employee_cases), Decimal("0")))])
        centers_ws = wb.create_sheet("Centros de costo")
        centers_ws.append(["codigo", "nombre", "presupuesto", "total_rendido", "total_aprobado", "diferencia", "cantidad_gastos", "cantidad_personas"])
        centers = sorted({str(e.get("cost_center", "") or "Sin centro de costo") for e in data["expenses"]})
        for center in centers:
            center_expenses = [e for e in data["expenses"] if str(e.get("cost_center", "") or "Sin centro de costo") == center]
            rendered = sum((_decimal(e.get("total_clp")) for e in center_expenses), Decimal("0"))
            approved = sum((_decimal(e.get("total_clp")) for e in center_expenses if _status(e.get("status")) == "approved"), Decimal("0"))
            people = {str(data["cases_by_id"].get(str(e.get("case_id", "")), {}).get("employee_phone", "")) for e in center_expenses}
            centers_ws.append(["", center, None, float(rendered), float(approved), None, len(center_expenses), len(people)])
        exceptions_ws = wb.create_sheet("Excepciones")
        exceptions_ws.append(["tipo", "gasto", "rendicion", "persona", "descripcion", "severidad", "accion_sugerida"])
        for warning in warnings:
            exceptions_ws.append([warning["type"], "", "", "", warning["description"], warning["severity"], "Revisar antes del cierre"])
        for sheet in wb.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
            for column in sheet.columns:
                letter = column[0].column_letter
                sheet.column_dimensions[letter].width = min(36, max(10, max(len(str(cell.value or "")) for cell in column) + 2))
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)) and any(token in str(sheet.cell(1, cell.column).value or "").lower() for token in ("total", "monto", "iva", "presupuesto", "diferencia", "saldo", "fondos")):
                        cell.number_format = '$#,##0'
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _csv(self, data: dict[str, Any]) -> bytes:
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=";", lineterminator="\n")
        writer.writerow(self._expense_headers())
        writer.writerows(self._expense_rows(data))
        return ("\ufeff" + buffer.getvalue()).encode("utf-8")

    def _zip(self, data: dict[str, Any], summary: dict[str, Any], export_id: str, user: dict[str, Any], generated_at: str, pdf_bytes: bytes, xlsx_bytes: bytes, csv_bytes: bytes) -> tuple[bytes, list[dict[str, str]], dict[str, Any]]:
        warnings: list[dict[str, str]] = []
        files: list[str] = []
        omitted: list[dict[str, str]] = []
        buffer = io.BytesIO()
        period = summary["period"]
        with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED, allowZip64=True) as archive:
            base_files = {f"Informe_Gastos_{period}.pdf": pdf_bytes, f"Detalle_Gastos_{period}.xlsx": xlsx_bytes}
            if csv_bytes:
                base_files[f"Gastos_{period}.csv"] = csv_bytes
            for name, content in base_files.items():
                archive.writestr(name, content)
                files.append(name)
            used: set[str] = set()
            for index, expense in enumerate(data["expenses"], 1):
                object_key = str(expense.get("receipt_object_key", "") or "")
                if not object_key:
                    continue
                case = data["cases_by_id"].get(str(expense.get("case_id", "")), {})
                employee = data["employees_by_phone"].get(str(case.get("employee_phone", case.get("phone", "")) or ""), {})
                path = self._receipt_zip_path(expense, case, employee, index=index)
                original = path
                collision = 2
                while path.lower() in used:
                    stem, dot, suffix = original.rpartition(".")
                    path = f"{stem}_{collision}.{suffix}" if dot else f"{original}_{collision}"
                    collision += 1
                used.add(path.lower())
                try:
                    content = self.storage_service.download_private_bytes(object_key=object_key)
                    archive.writestr(path, content)
                    files.append(path)
                except Exception as exc:
                    warning = {"type": "receipt_unavailable", "description": f"No se pudo incorporar comprobante de {expense.get('expense_id')}", "severity": "warning"}
                    warnings.append(warning)
                    omitted.append({"path": path, "reason": str(exc)})
            docs_by_case: dict[str, dict[str, Any]] = {}
            for document in data["documents"]:
                if document.get("object_key"):
                    docs_by_case[str(document.get("case_id", ""))] = document
            for case_id, document in docs_by_case.items():
                case = data["cases_by_id"].get(case_id, {})
                folder = _safe_name(case.get("context_label") or case_id, "rendicion")
                path = f"rendiciones/{folder}/resumen.pdf"
                try:
                    archive.writestr(path, self.storage_service.download_private_bytes(object_key=str(document["object_key"])))
                    files.append(path)
                except Exception as exc:
                    warnings.append({"type": "consolidated_pdf_unavailable", "description": f"No se pudo incorporar PDF consolidado de {case_id}", "severity": "warning"})
                    omitted.append({"path": path, "reason": str(exc)})
            manifest = {"export_id": export_id, "empresa": summary["company_name"], "periodo": period, "generado_en": generated_at, "generado_por": user.get("email", ""), "numero_rendiciones": summary["case_count"], "numero_gastos": summary["expense_count"], "total_clp": summary["total_clp"], "archivos_incluidos": files + ["manifest.json"], "archivos_omitidos": omitted, "errores_encontrados": warnings, "version_formato": FORMAT_VERSION}
            archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2).encode())
        if buffer.tell() > MAX_ZIP_BYTES:
            raise ValueError("El ZIP supera el límite de 250 MB")
        return buffer.getvalue(), warnings, manifest

    def _receipt_zip_path(self, expense: dict[str, Any], case: dict[str, Any], employee: dict[str, Any], index: int = 1) -> str:
        folder = _safe_name(f"{case.get('context_label') or case.get('case_id')}_{employee.get('name')}", "sin_rendicion")
        original = PurePosixPath(str(expense.get("receipt_object_key", "") or "comprobante")).name
        suffix = PurePosixPath(original).suffix[:10]
        merchant = _safe_name(expense.get("merchant"), "comprobante")
        return f"rendiciones/{folder}/comprobantes/{index:03d}_{merchant}_{_safe_name(expense.get('date'), 'sin_fecha')}{suffix}"

    @staticmethod
    def _format_clp(value: Any) -> str:
        return f"${int(value or 0):,}".replace(",", ".")

    @staticmethod
    def _public_row(row: dict[str, Any]) -> dict[str, Any]:
        public = {key: value for key, value in row.items() if not key.endswith("_object_key") and key not in {"snapshot_json"}}
        public["warnings"] = json_loads(row.get("warnings_json"), default=[])
        public["filters"] = json_loads(row.get("filters_json"), default={})
        public.pop("warnings_json", None)
        public.pop("filters_json", None)
        public["downloads"] = {
            fmt: bool(row.get(f"{fmt}_object_key"))
            for fmt in ("pdf", "xlsx", "csv", "zip")
        }
        return public
