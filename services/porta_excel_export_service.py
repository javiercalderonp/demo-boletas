from __future__ import annotations

import io
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SHEET_LAYOUTS = {
    "Gastos de Producción": {"total_row": 39, "amount_column": "D"},
    "Alimentación": {"total_row": 39, "amount_column": "D"},
    "Combustible": {"total_row": 39, "amount_column": "D"},
    "Estacionamientos-Peajes-Taxis": {"total_row": 39, "amount_column": "D"},
    "Boletas de Honorarios": {"total_row": 38, "amount_column": "F"},
    "Facturas": {"total_row": 39, "amount_column": "F"},
}
EXPECTED_SHEETS = ["Resumen", *SHEET_LAYOUTS]
DATA_START_ROW = 12
FORMAT_VERSION = "porta-2026-v1"


class PortaExcelExportService:
    """Render expenses into an untouched copy of Porta's original workbook."""

    def __init__(self, template_path: str) -> None:
        self.template_path = self._resolve_template_path(template_path)

    @staticmethod
    def _resolve_template_path(configured_path: str) -> Path:
        configured = Path(configured_path)
        if configured.is_file():
            return configured
        project_root = Path(__file__).resolve().parent.parent
        project_configured = project_root / configured_path
        if project_configured.is_file():
            return project_configured
        matches = list(project_root.glob("FORMATO*2026.xlsx"))
        if len(matches) == 1:
            return matches[0]
        raise FileNotFoundError("No se encontró la plantilla FORMATO RENDICIÓN 2026.xlsx")

    def render(self, expenses: list[dict[str, Any]]) -> bytes:
        workbook = load_workbook(self.template_path, data_only=False)
        if workbook.sheetnames != EXPECTED_SHEETS:
            raise ValueError("La plantilla Porta no contiene las siete hojas esperadas")

        grouped = {sheet_name: [] for sheet_name in SHEET_LAYOUTS}
        for expense in expenses:
            grouped[str(expense["porta_sheet"])].append(expense)

        total_rows: dict[str, int] = {}
        for sheet_name, layout in SHEET_LAYOUTS.items():
            worksheet = workbook[sheet_name]
            original_total_row = int(layout["total_row"])
            capacity = original_total_row - DATA_START_ROW
            required = len(grouped[sheet_name])
            inserted = max(0, required - capacity)
            if inserted:
                worksheet.insert_rows(original_total_row, amount=inserted)
                for row_number in range(original_total_row, original_total_row + inserted):
                    self._copy_data_row_style(worksheet, DATA_START_ROW, row_number)
            total_row = original_total_row + inserted
            total_rows[sheet_name] = total_row

            for row_number in range(DATA_START_ROW, total_row):
                for column in range(1, 7):
                    worksheet.cell(row_number, column).value = None
            for offset, expense in enumerate(grouped[sheet_name]):
                self._write_expense(worksheet, DATA_START_ROW + offset, sheet_name, expense)

            amount_column = str(layout["amount_column"])
            worksheet[f"{amount_column}{total_row}"] = (
                f"=SUM({amount_column}{DATA_START_ROW}:{amount_column}{total_row - 1})"
            )
            if sheet_name == "Boletas de Honorarios":
                worksheet[f"D{total_row}"] = f"=SUM(D{DATA_START_ROW}:D{total_row - 1})"
                worksheet[f"E{total_row}"] = f"=SUM(E{DATA_START_ROW}:E{total_row - 1})"
            elif sheet_name == "Facturas":
                worksheet[f"D{total_row}"] = f"=SUM(D{DATA_START_ROW}:D{total_row - 1})"
                worksheet[f"E{total_row}"] = f"=SUM(E{DATA_START_ROW}:E{total_row - 1})"

        summary = workbook["Resumen"]
        for cell in ("C4", "C5", "C6", "C7", "C9"):
            summary[cell] = None
        summary["C12"] = f"=+'Gastos de Producción'!D{total_rows['Gastos de Producción']}"
        summary["D12"] = f"='Gastos de Producción'!D{total_rows['Gastos de Producción']}"
        summary["C13"] = f"=+Alimentación!D{total_rows['Alimentación']}"
        summary["D13"] = f"=Alimentación!D{total_rows['Alimentación']}"
        summary["C14"] = f"=+Combustible!D{total_rows['Combustible']}"
        summary["D14"] = f"=Combustible!D{total_rows['Combustible']}"
        summary["C15"] = (
            f"=+'Estacionamientos-Peajes-Taxis'!D"
            f"{total_rows['Estacionamientos-Peajes-Taxis']}"
        )
        summary["D15"] = (
            f"='Estacionamientos-Peajes-Taxis'!D"
            f"{total_rows['Estacionamientos-Peajes-Taxis']}"
        )
        summary["C16"] = f"=+'Boletas de Honorarios'!D{total_rows['Boletas de Honorarios']}"
        summary["D16"] = f"=+'Boletas de Honorarios'!F{total_rows['Boletas de Honorarios']}"
        summary["C17"] = f"=+Facturas!F{total_rows['Facturas']}"
        summary["D17"] = f"=+Facturas!D{total_rows['Facturas']}"
        summary["C19"] = "=SUM(C12:C17)"
        summary["D19"] = "=SUM(D12:D17)"
        summary["C20"] = "=C9-C19"

        calculation = workbook.calculation
        calculation.fullCalcOnLoad = True
        calculation.forceFullCalc = True
        calculation.calcMode = "auto"
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _copy_data_row_style(worksheet, source_row: int, target_row: int) -> None:
        worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
        for column in range(1, worksheet.max_column + 1):
            source = worksheet.cell(source_row, column)
            target = worksheet.cell(target_row, column)
            target._style = copy(source._style)
            target.number_format = source.number_format
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)

    @staticmethod
    def _write_expense(worksheet, row: int, sheet_name: str, expense: dict[str, Any]) -> None:
        expense_date = expense.get("date")
        if isinstance(expense_date, str):
            try:
                expense_date = date.fromisoformat(expense_date[:10])
            except ValueError:
                expense_date = expense_date
        elif isinstance(expense_date, datetime):
            expense_date = expense_date.date()
        worksheet.cell(row, 1).value = expense_date
        worksheet.cell(row, 2).value = expense.get("document_number", "")
        worksheet.cell(row, 3).value = expense.get("detail", "")
        if sheet_name == "Facturas":
            worksheet.cell(row, 4).value = expense["net_amount"]
            worksheet.cell(row, 5).value = expense["tax_amount"]
            worksheet.cell(row, 6).value = expense["gross_amount"]
        elif sheet_name == "Boletas de Honorarios":
            worksheet.cell(row, 4).value = expense["net_amount"]
            worksheet.cell(row, 5).value = expense["withholding_amount"]
            worksheet.cell(row, 6).value = expense["gross_amount"]
        else:
            worksheet.cell(row, 4).value = expense["gross_amount"]
